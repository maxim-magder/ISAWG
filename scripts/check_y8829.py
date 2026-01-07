import openpyxl

filepath = 'ISOGG_dna-differences_and_other_info/Haplogroup Data 2019-2020-~/SNP Index_.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

# Find Y8829
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    name_cell = row[0].value
    if name_cell and 'Y8829' in str(name_cell):
        print(f'Row {row[0].row}:')
        print(f'  Name: {repr(name_cell)}')
        if len(row) > 3:
            print(f'  Alt names (col D): {repr(row[3].value)}')
        if len(row) > 2:
            print(f'  Col C: {repr(row[2].value)}')
        if len(row) > 1:
            print(f'  Col B: {repr(row[1].value)}')
        break
