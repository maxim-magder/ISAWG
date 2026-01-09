#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Haplogroup Data 2019-2020-~/2019-2020 Haplogroup E Tree.xlsx', data_only=True)
ws = wb.active

print(f"Total rows: {ws.max_row}")
print("\nSearching for M44...")

for row_idx in range(1, min(2000, ws.max_row + 1)):
    for col_idx in range(1, 20):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and 'M44' in str(val):
            print(f"\nFound 'M44' at row {row_idx}, col {col_idx}")
            print(f"Row {row_idx}:")
            for c in range(1, 10):
                v = ws.cell(row=row_idx, column=c).value
                if v:
                    print(f"  C{c}: {v}")
            
            # Check the row above (might be header)
            print(f"\nRow {row_idx-1} (potential header):")
            for c in range(1, 10):
                v = ws.cell(row=row_idx-1, column=c).value
                if v:
                    print(f"  C{c}: {v}")
            break
    else:
        continue
    break
