#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Haplogroup Data 2019-2020-~/2019-2020 Haplogroup E Tree.xlsx', data_only=True)
ws = wb['Sheet1']

print("Row 1327 (header row):")
for col_idx in range(1, 20):
    val = ws.cell(row=1327, column=col_idx).value
    if val:
        print(f"  C{col_idx}: {val}")

print("\nRow 1328 (first data row):")
for col_idx in range(1, 20):
    val = ws.cell(row=1328, column=col_idx).value
    if val:
        val_str = str(val)[:100]  # Limit length
        print(f"  C{col_idx}: {val_str}")

print("\nRow 1884 (M44 row):")
for col_idx in range(1, 20):
    val = ws.cell(row=1884, column=col_idx).value
    if val:
        val_str = str(val)[:200]  # Limit length
        print(f"  C{col_idx}: {val_str}")
