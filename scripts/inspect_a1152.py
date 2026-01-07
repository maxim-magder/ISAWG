import openpyxl

filepath = 'ISOGG_dna-differences_and_other_info/Haplogroup Data 2019-2020-~/SNP Index_.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

# Find A1152
for row in ws.iter_rows(min_row=2, max_row=min(200, ws.max_row)):
    if row[0].value == 'A1152':
        print(f'Found A1152 at row {row[0].row}')
        for i, cell in enumerate(row[:8]):
            print(f'  Col {i}: {repr(cell.value)}')
        break
