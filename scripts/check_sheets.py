#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Haplogroup Data 2019-2020-~/2019-2020 Haplogroup E Tree.xlsx', data_only=True)

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    ws = wb[sheet_name]
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    # Look for "Name" header
    for row_idx in range(1, min(100, ws.max_row + 1)):
        first_val = ws.cell(row=row_idx, column=1).value
        if first_val and str(first_val).strip().lower() == 'name':
            print(f"\n  Found 'Name' header at row {row_idx}!")
            for col_idx in range(1, 10):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    print(f"    C{col_idx}: {val}")
            break
