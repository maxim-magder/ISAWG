#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Haplogroup Data 2019-2020-~/2019-2020 Haplogroup E Tree.xlsx', data_only=True)
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
print("\nScanning for 'Name' header...")

for row_idx in range(1, min(2000, ws.max_row + 1)):
    first_cell = ws.cell(row=row_idx, column=1).value
    if first_cell:
        first_str = str(first_cell).strip()
        if first_str.lower() == 'name' or 'name' in first_str.lower():
            print(f"\nPossible header at row {row_idx}:")
            print(f"  C1: {first_cell}")
            for col_idx in range(2, 10):
                val = ws.cell(row=row_idx, column=col_idx).value
                print(f"  C{col_idx}: {val}")
