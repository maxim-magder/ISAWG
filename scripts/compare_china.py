#!/usr/bin/env python3
"""Compare main ISOGG data with China version to identify differences."""

import openpyxl
import json

def compare_haplogroup(letter):
    """Compare a specific haplogroup between main and China versions."""
    
    main_path = f'ISOGG_dna-differences_and_other_info/Haplogroup Data 2019-2020-~/2019-2020 Haplogroup {letter} Tree.xlsx'
    china_path = f'ISOGG_dna-differences_and_other_info/Y-DNA Haplogroup Tree 2019-2020 (for China users)/2019Haplogroup{letter}.xlsx'
    
    try:
        main_wb = openpyxl.load_workbook(main_path)
        china_wb = openpyxl.load_workbook(china_path)
    except Exception as e:
        print(f"Could not load {letter}: {e}")
        return None
    
    main_ws = main_wb.active
    china_ws = china_wb.active
    
    print(f"\n=== Haplogroup {letter} ===")
    print(f"Main:  {main_ws.dimensions}")
    print(f"China: {china_ws.dimensions}")
    
    # Extract haplogroup names from both
    def extract_haplogroups(ws, max_rows=5000):
        hgs = set()
        for row in range(1, min(max_rows, ws.max_row + 1)):
            for col in range(1, 10):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str):
                    val = val.strip()
                    # Check if looks like haplogroup
                    if len(val) >= 1 and val[0] in 'ABCDEFGHIJKLMNOPQRST':
                        if val[0] == letter or (letter in 'KP' and val[0] in 'KLMNOPQRST'):
                            hgs.add(val)
        return hgs
    
    main_hgs = extract_haplogroups(main_ws)
    china_hgs = extract_haplogroups(china_ws)
    
    only_main = main_hgs - china_hgs
    only_china = china_hgs - main_hgs
    
    print(f"Main haplogroups: {len(main_hgs)}")
    print(f"China haplogroups: {len(china_hgs)}")
    print(f"Only in Main: {len(only_main)}")
    print(f"Only in China: {len(only_china)}")
    
    if only_china:
        print(f"  China-only samples: {list(only_china)[:10]}")
    
    return {
        'main_count': len(main_hgs),
        'china_count': len(china_hgs),
        'only_main': len(only_main),
        'only_china': len(only_china)
    }


def main():
    results = {}
    for letter in 'ABCDEFGHIJKLMNOPQRST':
        result = compare_haplogroup(letter)
        if result:
            results[letter] = result
    
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    
    total_only_china = sum(r.get('only_china', 0) for r in results.values())
    total_only_main = sum(r.get('only_main', 0) for r in results.values())
    
    print(f"Total haplogroups only in China version: {total_only_china}")
    print(f"Total haplogroups only in Main version: {total_only_main}")


if __name__ == '__main__':
    main()
