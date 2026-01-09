#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Haplogroup Data 2019-2020-~/2019-2020 Haplogroup E Tree.xlsx', data_only=True)
ws = wb['Sheet1']

# Search for rs796742903 (M44's rs number)
print("Searching for M44's rs number: rs796742903")
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, 20):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and 'rs796742903' in str(val):
            print(f"\nFound at row {row_idx}, col {col_idx}")
            print(f"Full row {row_idx}:")
            for c in range(1, 10):
                v = ws.cell(row=row_idx, column=c).value
                if v:
                    print(f"  C{c}: {v}")
            
            # Check header
            print(f"\nRow {row_idx-1} (header?):")
            for c in range(1, 10):
                v = ws.cell(row=row_idx-1, column=c).value
                if v:
                    print(f"  C{c}: {v}")
            break
    else:
        continue
    break

# Also search for "rs #" header
print("\n" + "="*60)
print("Searching for 'rs #' header...")
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, 20):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and 'rs #' in str(val).lower():
            print(f"\nFound 'rs #' at row {row_idx}, col {col_idx}")
            print(f"Row {row_idx} headers:")
            for c in range(1, 10):
                v = ws.cell(row=row_idx, column=c).value
                if v:
                    print(f"  C{c}: {v}")
            break
    else:
        continue
    break
