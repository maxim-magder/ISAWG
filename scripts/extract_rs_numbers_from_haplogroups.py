#!/usr/bin/env python3
"""
Extract rs numbers from individual haplogroup XLSX files.
These files have more complete rs# data than the SNP Index.
"""

import openpyxl
import json
import os
from pathlib import Path

def parse_haplogroup_file(filepath):
    """Parse a single haplogroup XLSX file to extract SNP data with rs numbers."""
    print(f"\nProcessing: {os.path.basename(filepath)}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find header row (look for "rs #" column)
    # The SNP table is BELOW the tree, typically starting around row 1000-1500
    # Headers are in columns 11-17
    header_row = None
    header_col_offset = None
    
    for row_idx in range(1, ws.max_row + 1):
        # Scan across columns looking for "rs #"
        for col_idx in range(1, 30):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and 'rs #' in str(val).lower():
                # Found rs # column, check if nearby columns have Name and Haplogroup
                name_col = col_idx - 3  # Name should be 3 columns before rs #
                hg_col = col_idx - 2     # Haplogroup should be 2 columns before rs #
                
                name_val = ws.cell(row=row_idx, column=name_col).value
                hg_val = ws.cell(row=row_idx, column=hg_col).value
                
                if (name_val and str(name_val).strip().lower() == 'name' and
                    hg_val and 'haplogroup' in str(hg_val).lower()):
                    header_row = row_idx
                    header_col_offset = name_col - 1  # Offset from column 1
                    break
        
        if header_row:
            break
    
    if not header_row:
        print(f"  Warning: Could not find header row")
        return {}
    
    # Get column mapping (columns are offset by header_col_offset)
    col_map = {}
    for offset in range(0, 10):
        col_idx = header_col_offset + 1 + offset
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            col_name = str(val).strip().lower()
            col_map[col_name] = col_idx
    
    print(f"  Header at row {header_row}, cols start at {header_col_offset + 1}")
    print(f"  Headers: {list(col_map.keys())}")
    
    # Extract SNP data
    snp_data = {}
    rows_processed = 0
    name_col = col_map.get('name')
    
    if not name_col:
        print(f"  Warning: No 'name' column found")
        return {}
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Get SNP name
        name = ws.cell(row=row_idx, column=name_col).value
        if not name:
            continue
        
        name = str(name).strip()
        if not name or name.startswith('='):
            continue
        
        rows_processed += 1
        
        # Get rs number
        rs_col = col_map.get('rs #') or col_map.get('rs number') or col_map.get('rs numbers')
        rs_number = None
        if rs_col:
            rs_val = ws.cell(row=row_idx, column=rs_col).value
            if rs_val:
                rs_number = str(rs_val).strip()
                if rs_number in ['None', '', 'nan']:
                    rs_number = None
        
        # Get haplogroup
        hg_col = col_map.get('haplogroup') or col_map.get('subgroup name')
        haplogroup = None
        if hg_col:
            hg_val = ws.cell(row=row_idx, column=hg_col).value
            if hg_val:
                haplogroup = str(hg_val).strip()
        
        # Get build positions
        build37_col = col_map.get('build 37 #') or col_map.get('build 37 number')
        build38_col = col_map.get('build 38 #') or col_map.get('build 38 number')
        
        build37 = None
        build38 = None
        
        if build37_col:
            val = ws.cell(row=row_idx, column=build37_col).value
            if val:
                try:
                    build37 = int(float(val))
                except:
                    pass
        
        if build38_col:
            val = ws.cell(row=row_idx, column=build38_col).value
            if val:
                try:
                    build38 = int(float(val))
                except:
                    pass
        
        # Get mutation
        mut_col = col_map.get('mutation info') or col_map.get('mutation')
        mutation = None
        if mut_col:
            mut_val = ws.cell(row=row_idx, column=mut_col).value
            if mut_val:
                mutation = str(mut_val).strip()
        
        snp_data[name] = {
            'name': name,
            'haplogroup': haplogroup,
            'rs_number': rs_number,
            'build37_position': build37,
            'build38_position': build38,
            'mutation_info': mutation
        }
    
    print(f"  Processed {rows_processed} rows")
    with_rs = sum(1 for s in snp_data.values() if s['rs_number'])
    print(f"  Found {with_rs} SNPs with rs numbers")
    
    return snp_data


def main():
    """Main entry point."""
    source_dir = 'Haplogroup Data 2019-2020-~'
    
    if not os.path.exists(source_dir):
        print(f"Error: Directory not found: {source_dir}")
        return
    
    # Process all haplogroup files
    all_snp_data = {}
    
    haplogroup_files = sorted([
        f for f in os.listdir(source_dir)
        if f.endswith('.xlsx') and 'Haplogroup' in f and 'Tree' in f
    ])
    
    for filename in haplogroup_files:
        filepath = os.path.join(source_dir, filename)
        snp_data = parse_haplogroup_file(filepath)
        
        # Merge into all_snp_data (keep rs_number if new data has it)
        for name, data in snp_data.items():
            if name not in all_snp_data:
                all_snp_data[name] = data
            else:
                # Update rs_number if we found one
                if data['rs_number'] and not all_snp_data[name]['rs_number']:
                    all_snp_data[name]['rs_number'] = data['rs_number']
    
    # Save results
    output_file = 'haplogroup_rs_numbers.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_snp_data, f, indent=2, ensure_ascii=False)
    
    total_snps = len(all_snp_data)
    with_rs = sum(1 for s in all_snp_data.values() if s['rs_number'])
    
    print(f"\n{'='*60}")
    print(f"SUMMARY:")
    print(f"  Total SNPs: {total_snps:,}")
    print(f"  With rs numbers: {with_rs:,} ({100*with_rs/total_snps:.1f}%)")
    print(f"  Saved to: {output_file}")
    
    # Check M44 specifically
    if 'M44' in all_snp_data:
        print(f"\nM44 data:")
        print(json.dumps(all_snp_data['M44'], indent=2))


if __name__ == '__main__':
    main()
