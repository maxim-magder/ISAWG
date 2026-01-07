"""
ISOGG XLSX Data Processor

Parses Excel files from 2018-2020 containing:
- Haplogroup tree data (A-T)
- Tree Trunk
- SNP Index with mutation information

The XLSX files use column position to indicate tree depth:
- Column 1 = depth 0 (root)
- Column 2 = depth 1
- etc.

Each row typically has:
- Haplogroup name in one column (indicates depth)
- SNP list in the next column
"""

import json
import os
import re
from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class HaplogroupNode:
    """Represents a single haplogroup in the tree."""
    name: str
    snps: List[str] = field(default_factory=list)
    depth: int = 0
    parent_id: Optional[str] = None
    children: List[str] = field(default_factory=list)
    order_index: int = 0
    is_investigational: bool = False  # True if name ends with ~
    
    def to_dict(self) -> dict:
        return {
            'name': self.name,
            'snps': self.snps,
            'depth': self.depth,
            'parent_id': self.parent_id,
            'children': self.children,
            'order_index': self.order_index,
            'is_investigational': self.is_investigational
        }


@dataclass
class SNPEntry:
    """Represents a SNP with mutation information."""
    name: str
    haplogroup: str
    alternate_names: List[str] = field(default_factory=list)
    rs_number: Optional[str] = None
    build37_position: Optional[int] = None
    build38_position: Optional[int] = None
    mutation_info: Optional[str] = None
    
    def to_dict(self) -> dict:
        return {
            'name': self.name,
            'haplogroup': self.haplogroup,
            'alternate_names': self.alternate_names,
            'rs_number': self.rs_number,
            'build37_position': self.build37_position,
            'build38_position': self.build38_position,
            'mutation_info': self.mutation_info
        }


class XLSXTreeParser:
    """Parses haplogroup tree data from XLSX files."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.nodes: Dict[str, HaplogroupNode] = {}
        self.root_nodes: List[str] = []
        
    def parse(self) -> Dict[str, HaplogroupNode]:
        """Parse the XLSX file and return haplogroup nodes."""
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        
        # Find where tree data starts (look for 'Y' or 'Root' pattern)
        start_row = self._find_tree_start(ws)
        if start_row is None:
            print(f"Warning: Could not find tree start in {self.filepath}")
            return {}
        
        # Parse tree data
        order_index = 0
        parent_stack: List[Tuple[int, str]] = []  # (depth, name) stack
        
        for row_idx in range(start_row, ws.max_row + 1):
            result = self._parse_row(ws, row_idx)
            if result is None:
                continue
                
            name, snps, depth = result
            
            # Skip empty or whitespace-only names
            if not name or not name.strip():
                continue
            
            name = name.strip()
            
            # Check if investigational (ends with ~)
            is_investigational = name.endswith('~')
            
            # Create node
            node = HaplogroupNode(
                name=name,
                snps=snps,
                depth=depth,
                order_index=order_index,
                is_investigational=is_investigational
            )
            order_index += 1
            
            # Update parent stack and set parent
            while parent_stack and parent_stack[-1][0] >= depth:
                parent_stack.pop()
            
            if parent_stack:
                node.parent_id = parent_stack[-1][1]
                # Add as child to parent
                if node.parent_id in self.nodes:
                    self.nodes[node.parent_id].children.append(name)
            else:
                self.root_nodes.append(name)
            
            parent_stack.append((depth, name))
            self.nodes[name] = node
        
        return self.nodes
    
    def _find_tree_start(self, ws: Worksheet) -> Optional[int]:
        """Find the row where tree data starts."""
        for row_idx in range(1, min(100, ws.max_row + 1)):
            for col_idx in range(1, 10):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    val_clean = val.strip()
                    # Look for 'Y' as root or 'Root (Y-Adam)'
                    if val_clean == 'Y' or 'Root' in val_clean or val_clean == 'A0000':
                        # Check if next column or same row has SNP-like data
                        return row_idx
                    
                    # Also check for single-letter haplogroups (B, C, D, etc.)
                    # These files start with their root haplogroup
                    if len(val_clean) == 1 and val_clean in 'ABCDEFGHIJKLMNOPQRST':
                        # Verify next column has SNP-like data
                        next_val = ws.cell(row=row_idx, column=col_idx + 1).value
                        if next_val and isinstance(next_val, str):
                            # Check if it looks like SNPs (e.g., "M60, M181/Page32")
                            if ',' in next_val or '/' in next_val:
                                return row_idx
                    
                    # Also handle "B~" style entries (investigational root)
                    if len(val_clean) == 2 and val_clean[0] in 'ABCDEFGHIJKLMNOPQRST' and val_clean[1] == '~':
                        return row_idx
        return None
    
    def _parse_row(self, ws: Worksheet, row_idx: int) -> Optional[Tuple[str, List[str], int]]:
        """
        Parse a row to extract haplogroup name, SNPs, and depth.
        Returns (name, snps, depth) or None if row is empty/invalid.
        """
        name = None
        snps = []
        depth = 0
        
        # Scan columns to find haplogroup name (first non-empty cell that looks like a haplogroup)
        for col_idx in range(1, 30):  # Check up to column 30
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            
            val_str = str(val).strip()
            if not val_str:
                continue
            
            # Skip hyperlink formulas - they contain haplogroup names but we want the next column
            if val_str.startswith('=HYPERLINK'):
                # Extract name from hyperlink if needed, but usually next col has it
                continue
            
            # Skip metadata/header text
            if any(skip in val_str.lower() for skip in [
                'could you adopt', 'downloading', 'version', 'criteria',
                'contact', 'font colors', 'symbols', 'click on', 'e-mail',
                'haplogroup tree', 'subclades', 'main page', 'listing',
                'papers', 'glossary', 'copyright', 'links:', 'newly confirmed',
                'investigational', 'did you notice', 'just below',
                'et al', 'journal', 'abstract', 'doi:', 'genetics', 'phylogen',
                'chromosom', 'american', 'lineage', 'revised', 'origin', 'diversity'
            ]):
                continue
            
            # Skip special markers
            if val_str.startswith('[') and val_str.endswith(']'):
                continue
            
            # Check if this looks like a haplogroup name
            if self._is_haplogroup_name(val_str):
                name = val_str
                depth = col_idx - 1  # Column 1 = depth 0
                
                # Next column should have SNPs
                next_val = ws.cell(row=row_idx, column=col_idx + 1).value
                if next_val:
                    snps = self._parse_snps(str(next_val))
                break
            
            # Check if this looks like SNP list (comma separated with alphanumeric codes)
            if col_idx > 1 and ',' in val_str and self._looks_like_snp_list(val_str):
                # Previous column had the name
                prev_val = ws.cell(row=row_idx, column=col_idx - 1).value
                if prev_val and self._is_haplogroup_name(str(prev_val).strip()):
                    name = str(prev_val).strip()
                    depth = col_idx - 2
                    snps = self._parse_snps(val_str)
                    break
        
        if name:
            return (name, snps, depth)
        return None
    
    def _is_haplogroup_name(self, val: str) -> bool:
        """Check if a value looks like a haplogroup name."""
        if not val:
            return False
        
        # Clean up
        val = val.strip()
        
        # Standard haplogroup patterns: A, A1, A1b, A1b1a2, BT, CT, etc.
        # Also handles names like "Root (Y-Adam)", "Y"
        if val in ['Y', 'BT', 'CT', 'DE', 'CF', 'GHIJK', 'HIJK', 'IJK', 'IJ', 'LT', 'NO', 'NOP']:
            return True
        
        if 'Root' in val or 'Y-Adam' in val:
            return True
        
        # Pattern: Letter(s) followed by optional numbers/letters, ending with optional ~
        # Examples: A, A0, A00, A1b1, R1b1a1b1a1a2c1a, I2a1~
        pattern = r'^[A-Z][0-9a-zA-Z\-]*~?$'
        if re.match(pattern, val):
            return True
        
        return False
    
    def _looks_like_snp_list(self, val: str) -> bool:
        """Check if value looks like a comma-separated SNP list."""
        parts = val.split(',')
        if len(parts) < 2:
            return False
        
        # Check if parts look like SNP names
        snp_like = 0
        for part in parts[:5]:  # Check first 5
            part = part.strip()
            if re.match(r'^[A-Z0-9][A-Z0-9\.\-/]+$', part, re.IGNORECASE):
                snp_like += 1
        
        return snp_like >= 2
    
    def _parse_snps(self, snp_string: str) -> List[str]:
        """Parse a comma-separated string of SNPs."""
        if not snp_string:
            return []
        
        snps = []
        for part in snp_string.split(','):
            snp = part.strip()
            if snp:
                snps.append(snp)
        
        return snps
    
    def to_dict(self) -> dict:
        """Convert parsed data to dictionary format."""
        return {
            'nodes': {name: node.to_dict() for name, node in self.nodes.items()},
            'root_nodes': self.root_nodes
        }
    
    def to_json(self, filepath: str):
        """Save parsed data to JSON file."""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)


class XLSXSNPIndexParser:
    """Parses SNP Index data from XLSX files."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.snps: Dict[str, SNPEntry] = {}
    
    def parse(self) -> Dict[str, SNPEntry]:
        """Parse the SNP Index XLSX file."""
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        
        # Find header row (look for 'Name' or 'Subgroup Name')
        header_row = self._find_header_row(ws)
        if header_row is None:
            print(f"Warning: Could not find header in {self.filepath}")
            return {}
        
        # Get column mapping
        col_map = self._get_column_mapping(ws, header_row)
        
        # Parse data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            entry = self._parse_snp_row(ws, row_idx, col_map)
            if entry and entry.name:
                self.snps[entry.name] = entry
        
        return self.snps
    
    def _find_header_row(self, ws: Worksheet) -> Optional[int]:
        """Find the header row."""
        for row_idx in range(1, 20):
            for col_idx in range(1, 10):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    val_lower = val.lower().strip()
                    if val_lower in ['name', 'subgroup name']:
                        return row_idx
        return None
    
    def _get_column_mapping(self, ws: Worksheet, header_row: int) -> Dict[str, int]:
        """Get mapping of column names to indices."""
        col_map = {}
        for col_idx in range(1, 20):
            val = ws.cell(row=header_row, column=col_idx).value
            if val:
                val_lower = str(val).lower().strip()
                col_map[val_lower] = col_idx
        return col_map
    
    def _parse_snp_row(self, ws: Worksheet, row_idx: int, col_map: Dict[str, int]) -> Optional[SNPEntry]:
        """Parse a single SNP row."""
        def get_val(key: str) -> Any:
            if key in col_map:
                return ws.cell(row=row_idx, column=col_map[key]).value
            return None
        
        # Get SNP name - could be in 'name' column or second column
        name = get_val('name')
        if not name:
            # Try alternate positions
            for key in ['snp name', 'snp']:
                name = get_val(key)
                if name:
                    break
        
        if not name:
            return None
        
        name = str(name).strip()
        if not name or name.startswith('='):
            return None
        
        # Get haplogroup
        haplogroup = get_val('subgroup name') or get_val('haplogroup') or ''
        if haplogroup:
            haplogroup = str(haplogroup).strip()
        
        # Get alternate names - normalize separators and formatting
        alt_names_raw = get_val('alternate names') or get_val('other names') or ''
        alt_names = []
        if alt_names_raw:
            # Replace newlines, slashes, and semicolons with commas for consistent splitting
            normalized = str(alt_names_raw)
            normalized = normalized.replace('\n', ',').replace('/', ',').replace(';', ',')
            for part in normalized.split(','):
                part = part.strip()
                # Skip empty parts or notes like "to tree" "added"
                if part and not part.startswith('to ') and part not in ['added', 'tree']:
                    alt_names.append(part)
        
        # Get rs number
        rs_num = get_val('rs numbers') or get_val('rs #') or get_val('rs number')
        if rs_num:
            rs_num = str(rs_num).strip()
            if not rs_num or rs_num == 'None':
                rs_num = None
        
        # Get build positions
        build37 = get_val('build 37 number') or get_val('build 37 #')
        build38 = get_val('build 38 number') or get_val('build 38 #')
        
        try:
            build37 = int(float(build37)) if build37 else None
        except (ValueError, TypeError):
            build37 = None
        
        try:
            build38 = int(float(build38)) if build38 else None
        except (ValueError, TypeError):
            build38 = None
        
        # Get mutation info
        mutation = get_val('mutation info') or get_val('mutation')
        if mutation:
            mutation = str(mutation).strip()
        
        return SNPEntry(
            name=name,
            haplogroup=haplogroup,
            alternate_names=alt_names,
            rs_number=rs_num,
            build37_position=build37,
            build38_position=build38,
            mutation_info=mutation
        )
    
    def to_dict(self) -> dict:
        """Convert to dictionary."""
        return {name: snp.to_dict() for name, snp in self.snps.items()}
    
    def to_json(self, filepath: str):
        """Save to JSON file."""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)


class ConversionTableParser:
    """Parses haplogroup name conversion tables (TSV format)."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.conversions: Dict[str, Dict[str, str]] = {}  # from_year -> {old_name: new_name}
    
    def parse(self) -> Dict[str, Dict[str, str]]:
        """Parse the conversion table TSV file."""
        with open(self.filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        if not lines:
            return {}
        
        # Parse header to get year columns
        header = lines[0].strip().split('\t')
        
        # Build conversion mapping
        for line in lines[1:]:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                old_name = parts[0].strip()
                new_name = parts[1].strip()
                
                # Handle comma-separated multiple targets
                if new_name.startswith('"') and new_name.endswith('"'):
                    new_name = new_name[1:-1]  # Remove quotes
                
                # Use header years as key
                from_year = header[0].replace('ISOGG ', '') if header else '2007'
                to_year = header[1].replace('ISOGG ', '') if len(header) > 1 else '2008'
                
                key = f"{from_year}_to_{to_year}"
                if key not in self.conversions:
                    self.conversions[key] = {}
                
                self.conversions[key][old_name] = new_name
        
        return self.conversions
    
    def to_json(self, filepath: str):
        """Save to JSON file."""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.conversions, f, indent=2, ensure_ascii=False)


def process_year_xlsx(base_dir: str, year: str, output_dir: str):
    """Process all XLSX files for a given year."""
    
    # Map year to directory name
    year_dirs = {
        '2018': 'Haplogroup Data 2018',
        '2019-2020': 'Haplogroup Data 2019-2020-~',
        '2018-china': 'Y-DNA Haplogroup Tree 2018 - (for China users)',
        '2019-2020-china': 'Y-DNA Haplogroup Tree 2019-2020 (for China users)'
    }
    
    if year not in year_dirs:
        print(f"Unknown year: {year}")
        return
    
    year_dir = os.path.join(base_dir, year_dirs[year])
    if not os.path.exists(year_dir):
        print(f"Directory not found: {year_dir}")
        return
    
    # Create output directory
    year_output = os.path.join(output_dir, year)
    os.makedirs(year_output, exist_ok=True)
    
    # Haplogroup pattern for tree files
    hg_pattern = re.compile(r'(Haplogroup|2019-2020 Haplogroup|2019Haplogroup)\s*([A-T])\s*(Tree)?\.xlsx', re.IGNORECASE)
    trunk_pattern = re.compile(r'(Tree\s*Trunk|TreeTrunk|2019TreeTrunk)\.xlsx', re.IGNORECASE)
    snp_pattern = re.compile(r'SNP\s*Index', re.IGNORECASE)
    
    all_nodes = {}
    all_root_nodes = []
    
    # Process each file
    for filename in sorted(os.listdir(year_dir)):
        if not filename.endswith('.xlsx'):
            continue
        
        filepath = os.path.join(year_dir, filename)
        
        # Check file type
        hg_match = hg_pattern.search(filename)
        trunk_match = trunk_pattern.search(filename)
        snp_match = snp_pattern.search(filename)
        
        if hg_match:
            # Haplogroup tree file
            letter = hg_match.group(2).upper()
            print(f"  Processing Haplogroup {letter}: {filename}")
            
            parser = XLSXTreeParser(filepath)
            nodes = parser.parse()
            
            # Save individual haplogroup
            hg_output = os.path.join(year_output, 'individual_haplogroups')
            os.makedirs(hg_output, exist_ok=True)
            parser.to_json(os.path.join(hg_output, f'haplogroup_{letter}.json'))
            
            # Merge into combined
            all_nodes.update(nodes)
            all_root_nodes.extend(parser.root_nodes)
            
        elif trunk_match:
            print(f"  Processing Tree Trunk: {filename}")
            parser = XLSXTreeParser(filepath)
            nodes = parser.parse()
            parser.to_json(os.path.join(year_output, 'tree_trunk.json'))
            
        elif snp_match:
            print(f"  Processing SNP Index: {filename}")
            parser = XLSXSNPIndexParser(filepath)
            snps = parser.parse()
            parser.to_json(os.path.join(year_output, 'snp_index.json'))
            print(f"    Parsed {len(snps)} SNPs")
    
    # Save merged tree
    if all_nodes:
        merged = {
            'year': year,
            'nodes': {name: node.to_dict() if isinstance(node, HaplogroupNode) else node 
                      for name, node in all_nodes.items()},
            'root_nodes': list(set(all_root_nodes))
        }
        with open(os.path.join(year_output, 'tree.json'), 'w', encoding='utf-8') as f:
            json.dump(merged, f, indent=2, ensure_ascii=False)
        print(f"  Merged tree: {len(all_nodes)} nodes")


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Parse ISOGG XLSX files')
    parser.add_argument('--input-dir', '-i', 
                        default='ISOGG_dna-differences_and_other_info',
                        help='Input directory containing XLSX files')
    parser.add_argument('--output-dir', '-o',
                        default='output_xlsx',
                        help='Output directory for JSON files')
    parser.add_argument('--year', '-y',
                        choices=['2018', '2019-2020', '2018-china', '2019-2020-china', 'all'],
                        default='all',
                        help='Year to process')
    
    args = parser.parse_args()
    
    years_to_process = ['2018', '2019-2020', '2018-china', '2019-2020-china'] if args.year == 'all' else [args.year]
    
    os.makedirs(args.output_dir, exist_ok=True)
    
    for year in years_to_process:
        print(f"\nProcessing {year}...")
        process_year_xlsx(args.input_dir, year, args.output_dir)
    
    print("\nDone!")


if __name__ == '__main__':
    main()
